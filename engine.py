
from __future__ import annotations

import ast
import math
import operator as op
import re
from collections import defaultdict
from copy import deepcopy
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
ALUMINIUM_FILE = BASE_DIR / "ALUMINIUM OPTIMSATION CALCULATOR - TEMPLATE.xlsx"
GLASS_OFFCUT_FILE = BASE_DIR / "GLASS OFFCUT CHECKER.xlsx"
GLASS_OPTIMISER_FILE = BASE_DIR / "GLASS CUTTING OPTIMISER TEMPLATE.xlsx"

CELL_REF_RE = re.compile(r'(?<![A-Z0-9_])(\$?[A-Z]{1,3}\$?\d+)\b')
OPS = {
    ast.Add: op.add,
    ast.Sub: op.sub,
    ast.Mult: op.mul,
    ast.Div: op.truediv,
    ast.USub: op.neg,
    ast.UAdd: op.pos,
    ast.Pow: op.pow,
}


def safe_eval(expr: str) -> float:
    def _eval(node):
        if isinstance(node, ast.Expression):
            return _eval(node.body)
        if isinstance(node, ast.Constant):
            return node.value
        if isinstance(node, ast.Num):  # pragma: no cover
            return node.n
        if isinstance(node, ast.BinOp):
            return OPS[type(node.op)](_eval(node.left), _eval(node.right))
        if isinstance(node, ast.UnaryOp):
            return OPS[type(node.op)](_eval(node.operand))
        raise ValueError(f"Unsupported expression: {ast.dump(node)}")
    return float(_eval(ast.parse(expr, mode="eval")))


def _norm(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "").strip())


def _variant_name(inputs: List[str], idx: int) -> str:
    if inputs:
        key = " + ".join(i.title() for i in inputs)
    else:
        key = "Standard"
    return f"Variant {idx}: {key}"


class TypologyCatalog:
    def __init__(self, aluminium_path: Path = ALUMINIUM_FILE):
        self.aluminium_wb = openpyxl.load_workbook(aluminium_path, data_only=False)
        self.profile_ws = self.aluminium_wb["PROFILE SIZE CALCULATOR"]
        self.weight_ws = self.aluminium_wb["WEIGHT CATALOGUE"]
        self._typologies = self._parse_typologies()
        self._weights = self._load_weights()

    @property
    def typologies(self) -> Dict[str, Dict[str, Any]]:
        return self._typologies

    @property
    def weights(self) -> Dict[str, float]:
        return self._weights

    def _load_weights(self) -> Dict[str, float]:
        out = {}
        for row in range(2, self.weight_ws.max_row + 1):
            code = _norm(self.weight_ws[f"B{row}"].value)
            kgm = self.weight_ws[f"D{row}"].value
            if code and isinstance(kgm, (int, float)):
                out[code] = float(kgm)
        return out

    def _parse_typologies(self) -> Dict[str, Dict[str, Any]]:
        ws = self.profile_ws
        titles: List[Tuple[int, str]] = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(2, c).value
            if isinstance(val, str) and val.strip():
                titles.append((c, _norm(val)))

        parsed: Dict[str, Dict[str, Any]] = {}
        for idx, (start_col, title) in enumerate(titles):
            end_col = titles[idx + 1][0] - 1 if idx + 1 < len(titles) else ws.max_column
            block_starts = []
            for r in range(1, ws.max_row + 1):
                vals = [ws.cell(r, c).value for c in range(start_col, end_col + 1)]
                if any(isinstance(v, str) and _norm(v) in ("WINDOW SIZE", "OVERALL SIZE") for v in vals):
                    block_starts.append(r)
            block_starts.append(ws.max_row + 1)

            variants = []
            for bi in range(len(block_starts) - 1):
                r0 = block_starts[bi]
                r1 = block_starts[bi + 1]

                label_col = None
                for rr in range(r0, min(r0 + 4, r1)):
                    for c in range(start_col, end_col + 1):
                        v = _norm(ws.cell(rr, c).value)
                        if v in ("OVERALL WIDTH", "OVERALL HEIGHT"):
                            label_col = c
                            break
                    if label_col:
                        break
                if not label_col:
                    continue

                value_col = label_col + 1
                profile_header_row = None
                profile_col = None
                for rr in range(r0, r1):
                    for c in range(start_col, end_col + 1):
                        if _norm(ws.cell(rr, c).value) == "PROFILE" and _norm(ws.cell(rr, c + 1).value) == "SIZE":
                            profile_header_row = rr
                            profile_col = c
                            break
                    if profile_header_row:
                        break
                if not profile_header_row:
                    continue

                glass_col = None
                for c in range(label_col + 1, end_col + 1):
                    if _norm(ws.cell(r0, c).value) == "GLASS SIZES" and _norm(ws.cell(r0 + 1, c).value) == "WIDTH":
                        glass_col = c
                        break

                inputs = []
                for rr in range(r0 + 1, profile_header_row - 1):
                    label = _norm(ws.cell(rr, label_col).value)
                    value = ws.cell(rr, value_col).value
                    if label:
                        inputs.append({
                            "row": rr,
                            "label": label,
                            "cell": ws.cell(rr, value_col).coordinate,
                            "formula": value if isinstance(value, str) and value.startswith("=") else None,
                        })

                profiles = []
                for rr in range(profile_header_row + 1, r1):
                    prof = _norm(ws.cell(rr, profile_col).value)
                    if not prof:
                        continue
                    if prof in {"WINDOW SIZE", "OVERALL SIZE", "GLASS SIZES", "PROFILES REQUIRED", "PROFILE"}:
                        continue
                    profiles.append({
                        "row": rr,
                        "profile": prof,
                        "size_cell": ws.cell(rr, profile_col + 1).coordinate,
                        "qty_cell": ws.cell(rr, profile_col + 2).coordinate,
                        "cut_cell": ws.cell(rr, profile_col + 3).coordinate,
                    })

                glass_rows = []
                if glass_col:
                    for rr in range(r0 + 2, profile_header_row - 1):
                        if any(ws.cell(rr, glass_col + offset).value is not None for offset in range(3)):
                            glass_rows.append({
                                "row": rr,
                                "width_cell": ws.cell(rr, glass_col).coordinate,
                                "height_cell": ws.cell(rr, glass_col + 1).coordinate,
                                "qty_cell": ws.cell(rr, glass_col + 2).coordinate,
                            })

                input_labels = [item["label"] for item in inputs if item["formula"] is None]
                variants.append({
                    "variant_index": len(variants) + 1,
                    "variant_name": _variant_name(input_labels, len(variants) + 1),
                    "title": title,
                    "key": f"{title}__{len(variants)+1}",
                    "block_start_row": r0,
                    "input_labels": input_labels,
                    "inputs": inputs,
                    "profiles": profiles,
                    "glass": glass_rows,
                    "value_col": value_col,
                    "profile_col": profile_col,
                    "glass_col": glass_col,
                })
            if variants:
                parsed[title] = {"title": title, "variants": variants}
        return parsed

    def variant_lookup(self) -> Dict[str, Dict[str, Any]]:
        out = {}
        for title, payload in self.typologies.items():
            for variant in payload["variants"]:
                out[variant["key"]] = variant
        return out

    def list_variant_options(self) -> List[Tuple[str, str]]:
        opts = []
        for title, payload in self.typologies.items():
            for variant in payload["variants"]:
                label = f"{title} · {variant['variant_name']}"
                opts.append((variant["key"], label))
        return opts

    def evaluate_variant(self, variant_key: str, inputs: Dict[str, float]) -> Dict[str, Any]:
        ws = self.profile_ws
        variant = self.variant_lookup()[variant_key]
        cache: Dict[str, float] = {}

        for item in variant["inputs"]:
            if item["formula"] is None:
                cache[item["cell"]] = float(inputs.get(item["label"], 0) or 0)

        def cell_value(coord: str) -> float:
            coord = coord.replace("$", "")
            if coord in cache:
                return cache[coord]
            val = ws[coord].value
            if isinstance(val, str) and val.startswith("="):
                expr = val[1:]
                expr = CELL_REF_RE.sub(lambda m: str(cell_value(m.group(1))), expr)
                cache[coord] = safe_eval(expr)
            elif val is None:
                cache[coord] = 0.0
            else:
                cache[coord] = float(val)
            return cache[coord]

        derived_inputs = {}
        for item in variant["inputs"]:
            derived_inputs[item["label"]] = cell_value(item["cell"])

        profiles = []
        for row in variant["profiles"]:
            size = round(cell_value(row["size_cell"]), 3)
            qty = round(cell_value(row["qty_cell"]), 3)
            cut = round(cell_value(row["cut_cell"]), 3)
            profiles.append({
                "profile": row["profile"],
                "length_mm": size,
                "qty": qty,
                "cut_degree": cut,
            })

        glass = []
        for row in variant["glass"]:
            width = round(cell_value(row["width_cell"]), 3)
            height = round(cell_value(row["height_cell"]), 3)
            qty = round(cell_value(row["qty_cell"]), 3)
            glass.append({
                "width_mm": width,
                "height_mm": height,
                "qty": qty,
            })

        return {
            "variant": variant,
            "resolved_inputs": derived_inputs,
            "profiles": profiles,
            "glass": glass,
        }


def load_default_glass_offcuts(glass_offcut_path: Path = GLASS_OFFCUT_FILE) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(glass_offcut_path, data_only=True)
    ws = wb["GLASS OFFCUTS LIST"]
    items = []
    for row in range(2, ws.max_row + 1):
        width = ws[f"A{row}"].value
        height = ws[f"B{row}"].value
        spec = _norm(ws[f"C{row}"].value)
        if isinstance(width, (int, float)) and isinstance(height, (int, float)) and spec:
            items.append({
                "id": f"G{row-1}",
                "spec": spec,
                "width_mm": float(width),
                "height_mm": float(height),
                "qty": 1,
            })
    return items


def load_glass_specs() -> List[str]:
    wb = openpyxl.load_workbook(GLASS_OFFCUT_FILE, data_only=True)
    ws = wb["GLASS SPECIFICATIONS"]
    specs = []
    for row in range(2, ws.max_row + 1):
        val = _norm(ws[f"A{row}"].value)
        if val:
            specs.append(val)
    return specs


def default_aluminium_offcuts() -> List[Dict[str, Any]]:
    return []


def expand_window_rows(
    windows: List[Dict[str, Any]],
    catalog: TypologyCatalog,
    glass_spec: str,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[str]]:
    profile_rows = []
    glass_rows = []
    warnings = []

    for idx, row in enumerate(windows, start=1):
        variant_key = row.get("variant_key")
        if not variant_key:
            continue
        qty_windows = int(row.get("window_qty", 1) or 1)
        eval_inputs = {k: float(row.get(k, 0) or 0) for k in (
            "OVERALL WIDTH",
            "OVERALL HEIGHT",
            "VENT WIDTH",
            "BOTTOM FIXED HEIGHT",
            "BOTTOM CLEARANCE REQUIRED",
            "MAIN VENT WIDTH",
        )}

        try:
            evaluated = catalog.evaluate_variant(variant_key, eval_inputs)
        except Exception as exc:
            warnings.append(f"Row {idx} could not be calculated: {exc}")
            continue

        for p in evaluated["profiles"]:
            length = float(p["length_mm"])
            qty = int(round(float(p["qty"]) * qty_windows))
            if length <= 0 or qty <= 0:
                warnings.append(f"{row.get('label','W'+str(idx))}: skipped invalid profile row {p['profile']} ({length} mm, qty {qty}).")
                continue
            profile_rows.append({
                "window_label": row.get("label", f"W{idx}"),
                "window_type": evaluated["variant"]["title"],
                "variant_name": evaluated["variant"]["variant_name"],
                "profile": p["profile"],
                "length_mm": length,
                "qty": qty,
                "cut_degree": p["cut_degree"],
            })

        gindex = 1
        for g in evaluated["glass"]:
            width = float(g["width_mm"])
            height = float(g["height_mm"])
            qty = int(round(float(g["qty"]) * qty_windows))
            if width <= 0 or height <= 0 or qty <= 0:
                warnings.append(f"{row.get('label','W'+str(idx))}: skipped invalid glass row ({width} x {height}, qty {qty}).")
                continue
            for i in range(qty):
                glass_rows.append({
                    "piece_id": f"{row.get('label', f'W{idx}')}-G{gindex}",
                    "window_label": row.get("label", f"W{idx}"),
                    "window_type": evaluated["variant"]["title"],
                    "variant_name": evaluated["variant"]["variant_name"],
                    "spec": row.get("glass_spec") or glass_spec,
                    "width_mm": width,
                    "height_mm": height,
                    "rotatable": True,
                })
                gindex += 1

    return profile_rows, glass_rows, warnings


def _expand_aluminium_offcuts(raw_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    expanded = []
    for row in raw_rows:
        qty = int(row.get("qty", 1) or 1)
        for i in range(qty):
            expanded.append({
                "id": f"{row.get('profile','AL')}-{len(expanded)+1}",
                "profile": _norm(row.get("profile")),
                "length_mm": float(row.get("length_mm", 0) or 0),
            })
    return expanded


def _expand_glass_offcuts(raw_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    expanded = []
    for row in raw_rows:
        qty = int(row.get("qty", 1) or 1)
        for _ in range(qty):
            expanded.append({
                "id": f"{row.get('id') or 'G'}-{len(expanded)+1}",
                "spec": _norm(row.get("spec")),
                "width_mm": float(row.get("width_mm", 0) or 0),
                "height_mm": float(row.get("height_mm", 0) or 0),
            })
    return expanded


def optimise_aluminium(
    profile_rows: List[Dict[str, Any]],
    stock_length_mm: float,
    kerf_mm: float,
    raw_offcuts: List[Dict[str, Any]],
) -> Dict[str, Any]:
    cuts = []
    for row in profile_rows:
        for _ in range(int(row["qty"])):
            cuts.append({
                "window_label": row["window_label"],
                "window_type": row["window_type"],
                "profile": row["profile"],
                "length_mm": float(row["length_mm"]),
                "cut_degree": row["cut_degree"],
            })

    offcuts = _expand_aluminium_offcuts(raw_offcuts)
    offcut_jobs = []
    unmatched = []
    remaining_offcuts = deepcopy(offcuts)

    # Best-fit offcut allocation
    for cut in sorted(cuts, key=lambda x: x["length_mm"], reverse=True):
        candidates = [
            o for o in remaining_offcuts
            if o["profile"] == cut["profile"] and o["length_mm"] >= cut["length_mm"]
        ]
        if candidates:
            chosen = min(candidates, key=lambda x: x["length_mm"])
            remaining_before = chosen["length_mm"]
            chosen["length_mm"] = max(0.0, chosen["length_mm"] - cut["length_mm"] - kerf_mm)
            offcut_jobs.append({
                **cut,
                "source_offcut_id": chosen["id"],
                "source_length_mm": remaining_before,
                "remaining_after_mm": chosen["length_mm"],
            })
        else:
            unmatched.append(cut)

    bars = []
    grouped = defaultdict(list)
    for cut in unmatched:
        grouped[cut["profile"]].append(cut)

    for profile, items in grouped.items():
        items = sorted(items, key=lambda x: x["length_mm"], reverse=True)
        profile_bars = []
        for item in items:
            placed = False
            for bar in profile_bars:
                extra = kerf_mm if bar["cuts"] else 0
                if bar["used_mm"] + item["length_mm"] + extra <= stock_length_mm:
                    bar["cuts"].append(item)
                    bar["used_mm"] += item["length_mm"] + extra
                    placed = True
                    break
            if not placed:
                profile_bars.append({
                    "profile": profile,
                    "stock_length_mm": stock_length_mm,
                    "used_mm": item["length_mm"],
                    "cuts": [item],
                })
        bars.extend(profile_bars)

    new_stock_jobs = []
    bar_rows = []
    for bar_index, bar in enumerate(bars, start=1):
        x = 0.0
        for cut_index, cut in enumerate(bar["cuts"], start=1):
            start_x = x
            end_x = x + cut["length_mm"]
            x = end_x + kerf_mm
            row = {
                **cut,
                "bar_no": bar_index,
                "stock_length_mm": bar["stock_length_mm"],
                "start_mm": start_x,
                "end_mm": end_x,
            }
            new_stock_jobs.append(row)
        bar_rows.append({
            "bar_no": bar_index,
            "profile": bar["profile"],
            "stock_length_mm": bar["stock_length_mm"],
            "used_mm": bar["used_mm"],
            "waste_mm": max(0.0, bar["stock_length_mm"] - bar["used_mm"]),
            "cuts_count": len(bar["cuts"]),
        })

    return {
        "offcut_jobs": offcut_jobs,
        "new_stock_jobs": new_stock_jobs,
        "bars": bar_rows,
        "remaining_offcuts": remaining_offcuts,
        "total_bars": len(bar_rows),
        "total_waste_mm": sum(row["waste_mm"] for row in bar_rows),
    }


def _fits_rect(piece_w: float, piece_h: float, stock_w: float, stock_h: float, rotatable: bool):
    if piece_w <= stock_w and piece_h <= stock_h:
        return True, False
    if rotatable and piece_h <= stock_w and piece_w <= stock_h:
        return True, True
    return False, False


def optimise_glass(
    glass_rows: List[Dict[str, Any]],
    sheet_width_mm: float,
    sheet_height_mm: float,
    kerf_mm: float,
    raw_glass_offcuts: List[Dict[str, Any]],
) -> Dict[str, Any]:
    offcuts = _expand_glass_offcuts(raw_glass_offcuts)
    pieces = sorted(glass_rows, key=lambda x: x["width_mm"] * x["height_mm"], reverse=True)

    offcut_jobs = []
    leftover_offcuts = deepcopy(offcuts)
    remaining_pieces = []

    for piece in pieces:
        candidates = []
        for offcut in leftover_offcuts:
            if offcut["spec"] != piece["spec"]:
                continue
            fits, rotated = _fits_rect(piece["width_mm"], piece["height_mm"], offcut["width_mm"], offcut["height_mm"], piece.get("rotatable", True))
            if fits:
                waste = offcut["width_mm"] * offcut["height_mm"] - piece["width_mm"] * piece["height_mm"]
                candidates.append((waste, offcut, rotated))
        if candidates:
            _, chosen, rotated = min(candidates, key=lambda x: x[0])
            pw = piece["height_mm"] if rotated else piece["width_mm"]
            ph = piece["width_mm"] if rotated else piece["height_mm"]
            source_w = chosen["width_mm"]
            source_h = chosen["height_mm"]
            right_w = max(0.0, source_w - pw - kerf_mm)
            top_h = max(0.0, source_h - ph - kerf_mm)
            # simple guillotine split
            chosen["width_mm"] = right_w
            chosen["height_mm"] = ph
            if top_h > 0 and source_w > 0:
                leftover_offcuts.append({
                    "id": f"{chosen['id']}-T",
                    "spec": chosen["spec"],
                    "width_mm": source_w,
                    "height_mm": top_h,
                })
            offcut_jobs.append({
                **piece,
                "source_offcut_id": chosen["id"],
                "source_width_mm": source_w,
                "source_height_mm": source_h,
                "rotated": rotated,
            })
        else:
            remaining_pieces.append(piece)

    sheets = []
    optimiser_jobs = []
    unplaced_jobs = []
    grouped = defaultdict(list)
    for piece in remaining_pieces:
        grouped[piece["spec"]].append(piece)

    for spec, items in grouped.items():
        spec_sheets = []
        for piece in items:
            placed = False
            for sheet in spec_sheets:
                for shelf in sheet["shelves"]:
                    fits, rotated = _fits_rect(piece["width_mm"], piece["height_mm"], shelf["remaining_width"], shelf["height"], piece.get("rotatable", True))
                    if fits:
                        pw = piece["height_mm"] if rotated else piece["width_mm"]
                        ph = piece["width_mm"] if rotated else piece["height_mm"]
                        x = shelf["x"]
                        y = shelf["y"]
                        shelf["placements"].append((piece, x, y, pw, ph, rotated))
                        shelf["x"] += pw + kerf_mm
                        shelf["remaining_width"] -= pw + kerf_mm
                        placed = True
                        break
                if placed:
                    break

                # new shelf
                used_height = sum(sh["height"] + kerf_mm for sh in sheet["shelves"])
                fits, rotated = _fits_rect(piece["width_mm"], piece["height_mm"], sheet_width_mm, sheet_height_mm - used_height, piece.get("rotatable", True))
                if fits:
                    pw = piece["height_mm"] if rotated else piece["width_mm"]
                    ph = piece["width_mm"] if rotated else piece["height_mm"]
                    shelf = {
                        "y": used_height,
                        "x": pw + kerf_mm,
                        "height": ph,
                        "remaining_width": sheet_width_mm - pw - kerf_mm,
                        "placements": [(piece, 0.0, used_height, pw, ph, rotated)],
                    }
                    sheet["shelves"].append(shelf)
                    placed = True
                    break
            if not placed:
                fits, rotated = _fits_rect(piece["width_mm"], piece["height_mm"], sheet_width_mm, sheet_height_mm, piece.get("rotatable", True))
                if not fits:
                    unplaced_jobs.append({
                        **piece,
                        "reason": f"Piece does not fit on sheet {int(sheet_width_mm)} x {int(sheet_height_mm)} mm",
                        "sheet_width_mm": sheet_width_mm,
                        "sheet_height_mm": sheet_height_mm,
                    })
                    continue

                pw = piece["height_mm"] if rotated else piece["width_mm"]
                ph = piece["width_mm"] if rotated else piece["height_mm"]
                spec_sheets.append({
                    "spec": spec,
                    "sheet_no": len(spec_sheets) + 1,
                    "shelves": [{
                        "y": 0.0,
                        "x": pw + kerf_mm,
                        "height": ph,
                        "remaining_width": max(0.0, sheet_width_mm - pw - kerf_mm),
                        "placements": [(piece, 0.0, 0.0, pw, ph, rotated)],
                    }]
                })
        sheets.extend(spec_sheets)

    sheet_rows = []
    for global_sheet_no, sheet in enumerate(sheets, start=1):
        used_area = 0.0
        for shelf in sheet["shelves"]:
            for piece, x, y, pw, ph, rotated in shelf["placements"]:
                used_area += pw * ph
                optimiser_jobs.append({
                    **piece,
                    "sheet_no": global_sheet_no,
                    "spec": sheet["spec"],
                    "x_mm": x,
                    "y_mm": y,
                    "placed_width_mm": pw,
                    "placed_height_mm": ph,
                    "rotated": rotated,
                    "sheet_width_mm": sheet_width_mm,
                    "sheet_height_mm": sheet_height_mm,
                })
        sheet_area = sheet_width_mm * sheet_height_mm
        sheet_rows.append({
            "sheet_no": global_sheet_no,
            "spec": sheet["spec"],
            "sheet_width_mm": sheet_width_mm,
            "sheet_height_mm": sheet_height_mm,
            "used_area_mm2": used_area,
            "waste_area_mm2": max(0.0, sheet_area - used_area),
            "utilisation_pct": (used_area / sheet_area * 100.0) if sheet_area else 0.0,
        })

    return {
        "offcut_jobs": offcut_jobs,
        "optimiser_jobs": optimiser_jobs,
        "unplaced_jobs": unplaced_jobs,
        "sheets": sheet_rows,
        "remaining_offcuts": leftover_offcuts,
    }


def build_summary(
    windows: List[Dict[str, Any]],
    profile_rows: List[Dict[str, Any]],
    glass_rows: List[Dict[str, Any]],
    aluminium: Dict[str, Any],
    glass: Dict[str, Any],
    weights: Dict[str, float],
) -> Dict[str, Any]:
    total_profile_length_mm = sum(row["length_mm"] * row["qty"] for row in profile_rows)
    total_glass_area_mm2 = sum(row["width_mm"] * row["height_mm"] for row in glass_rows)

    kg_total = 0.0
    by_profile = defaultdict(lambda: {"cuts": 0, "length_mm": 0.0, "kg": 0.0})
    for row in profile_rows:
        by_profile[row["profile"]]["cuts"] += row["qty"]
        by_profile[row["profile"]]["length_mm"] += row["length_mm"] * row["qty"]

    for profile, payload in by_profile.items():
        code = None
        parts = profile.split()
        if parts:
            code = parts[0]
        kgm = weights.get(profile) or weights.get(code or "") or 0.0
        payload["kg"] = kgm * payload["length_mm"] / 1000.0
        kg_total += payload["kg"]

    return {
        "window_count": sum(int(w.get("window_qty", 1) or 1) for w in windows),
        "window_lines": len(windows),
        "profile_cuts": sum(row["qty"] for row in profile_rows),
        "glass_pieces": len(glass_rows),
        "total_profile_length_mm": total_profile_length_mm,
        "total_glass_area_m2": total_glass_area_mm2 / 1_000_000.0,
        "estimated_weight_kg": kg_total,
        "aluminium_offcut_hits": len(aluminium["offcut_jobs"]),
        "aluminium_new_bars": aluminium["total_bars"],
        "glass_offcut_hits": len(glass["offcut_jobs"]),
        "glass_new_sheets": len(glass["sheets"]),
        "profile_breakdown": by_profile,
    }


def _style_header(ws, row: int, start_col: int, end_col: int, fill="1F4E78"):
    thin = Side(style="thin", color="D9E2F3")
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)


def _autosize(ws):
    widths = defaultdict(int)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                widths[cell.column] = max(widths[cell.column], len(str(cell.value)))
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), 35)


def export_project_workbook(
    project_meta: Dict[str, Any],
    windows: List[Dict[str, Any]],
    profile_rows: List[Dict[str, Any]],
    glass_rows: List[Dict[str, Any]],
    aluminium: Dict[str, Any],
    glass: Dict[str, Any],
    summary: Dict[str, Any],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Aluminium Fabrication Project"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A3"] = "Project Name"
    ws["B3"] = project_meta.get("project_name")
    ws["A4"] = "Client"
    ws["B4"] = project_meta.get("client_name")
    ws["A5"] = "Finish"
    ws["B5"] = project_meta.get("finish")
    ws["A6"] = "Aluminium Stock Length (mm)"
    ws["B6"] = project_meta.get("stock_length_mm")
    ws["A7"] = "Glass Sheet Size (mm)"
    ws["B7"] = f"{project_meta.get('glass_sheet_width_mm')} x {project_meta.get('glass_sheet_height_mm')}"

    metrics = [
        ("Window lines", summary["window_lines"]),
        ("Total windows", summary["window_count"]),
        ("Profile cuts", summary["profile_cuts"]),
        ("Glass pieces", summary["glass_pieces"]),
        ("Total profile length (mm)", round(summary["total_profile_length_mm"], 1)),
        ("Total glass area (m²)", round(summary["total_glass_area_m2"], 3)),
        ("Estimated aluminium weight (kg)", round(summary["estimated_weight_kg"], 2)),
        ("Aluminium offcut hits", summary["aluminium_offcut_hits"]),
        ("New aluminium bars", summary["aluminium_new_bars"]),
        ("Glass offcut hits", summary["glass_offcut_hits"]),
        ("New glass sheets", summary["glass_new_sheets"]),
    ]
    start = 10
    for i, (label, value) in enumerate(metrics, start=start):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value

    # window schedule
    ws2 = wb.create_sheet("Window Schedule")
    headers = ["Label", "Typology Variant", "Window Qty", "Glass Spec", "OVERALL WIDTH", "OVERALL HEIGHT", "VENT WIDTH", "BOTTOM FIXED HEIGHT", "BOTTOM CLEARANCE REQUIRED", "MAIN VENT WIDTH"]
    for idx, h in enumerate(headers, start=1):
        ws2.cell(1, idx).value = h
    _style_header(ws2, 1, 1, len(headers))
    for r, row in enumerate(windows, start=2):
        ws2.cell(r, 1).value = row.get("label")
        ws2.cell(r, 2).value = row.get("variant_label")
        ws2.cell(r, 3).value = row.get("window_qty")
        ws2.cell(r, 4).value = row.get("glass_spec")
        for c, key in enumerate(headers[4:], start=5):
            ws2.cell(r, c).value = row.get(key)

    # profile pieces
    ws3 = wb.create_sheet("Profile Pieces")
    headers = ["Window", "Window Type", "Variant", "Profile", "Length (mm)", "Qty", "Cut Degree"]
    for i, h in enumerate(headers, start=1): ws3.cell(1, i).value = h
    _style_header(ws3, 1, 1, len(headers))
    for r, row in enumerate(profile_rows, start=2):
        vals = [row["window_label"], row["window_type"], row["variant_name"], row["profile"], row["length_mm"], row["qty"], row["cut_degree"]]
        for c, val in enumerate(vals, start=1): ws3.cell(r, c).value = val

    # aluminium offcuts
    ws4 = wb.create_sheet("Al Offcut Jobs")
    headers = ["Window", "Profile", "Cut (mm)", "Source Offcut", "Source Length", "Remaining After"]
    for i, h in enumerate(headers, start=1): ws4.cell(1, i).value = h
    _style_header(ws4, 1, 1, len(headers))
    for r, row in enumerate(aluminium["offcut_jobs"], start=2):
        vals = [row["window_label"], row["profile"], row["length_mm"], row["source_offcut_id"], row["source_length_mm"], row["remaining_after_mm"]]
        for c, val in enumerate(vals, start=1): ws4.cell(r, c).value = val

    ws5 = wb.create_sheet("Al New Stock Plan")
    headers = ["Bar No", "Window", "Profile", "Cut (mm)", "Start", "End", "Stock Length"]
    for i, h in enumerate(headers, start=1): ws5.cell(1, i).value = h
    _style_header(ws5, 1, 1, len(headers))
    for r, row in enumerate(aluminium["new_stock_jobs"], start=2):
        vals = [row["bar_no"], row["window_label"], row["profile"], row["length_mm"], row["start_mm"], row["end_mm"], row["stock_length_mm"]]
        for c, val in enumerate(vals, start=1): ws5.cell(r, c).value = val

    ws6 = wb.create_sheet("Glass Pieces")
    headers = ["Piece ID", "Window", "Spec", "Width (mm)", "Height (mm)"]
    for i, h in enumerate(headers, start=1): ws6.cell(1, i).value = h
    _style_header(ws6, 1, 1, len(headers))
    for r, row in enumerate(glass_rows, start=2):
        vals = [row["piece_id"], row["window_label"], row["spec"], row["width_mm"], row["height_mm"]]
        for c, val in enumerate(vals, start=1): ws6.cell(r, c).value = val

    ws7 = wb.create_sheet("Glass Offcut Jobs")
    headers = ["Piece ID", "Window", "Spec", "Width", "Height", "Source Offcut", "Rotated"]
    for i, h in enumerate(headers, start=1): ws7.cell(1, i).value = h
    _style_header(ws7, 1, 1, len(headers))
    for r, row in enumerate(glass["offcut_jobs"], start=2):
        vals = [row["piece_id"], row["window_label"], row["spec"], row["width_mm"], row["height_mm"], row["source_offcut_id"], row["rotated"]]
        for c, val in enumerate(vals, start=1): ws7.cell(r, c).value = val

    ws8 = wb.create_sheet("Glass Sheet Plan")
    headers = ["Piece ID", "Window", "Spec", "Sheet No", "X", "Y", "Placed Width", "Placed Height", "Rotated"]
    for i, h in enumerate(headers, start=1): ws8.cell(1, i).value = h
    _style_header(ws8, 1, 1, len(headers))
    for r, row in enumerate(glass["optimiser_jobs"], start=2):
        vals = [row["piece_id"], row["window_label"], row["spec"], row["sheet_no"], row["x_mm"], row["y_mm"], row["placed_width_mm"], row["placed_height_mm"], row["rotated"]]
        for c, val in enumerate(vals, start=1): ws8.cell(r, c).value = val

    ws9 = wb.create_sheet("Glass Sheets")
    headers = ["Sheet No", "Spec", "Sheet Width", "Sheet Height", "Used Area", "Waste Area", "Utilisation %"]
    for i, h in enumerate(headers, start=1): ws9.cell(1, i).value = h
    _style_header(ws9, 1, 1, len(headers))
    for r, row in enumerate(glass["sheets"], start=2):
        vals = [row["sheet_no"], row["spec"], row["sheet_width_mm"], row["sheet_height_mm"], row["used_area_mm2"], row["waste_area_mm2"], row["utilisation_pct"] / 100]
        for c, val in enumerate(vals, start=1): ws9.cell(r, c).value = val
        ws9.cell(r, 7).number_format = "0.0%"

    for sheet in wb.worksheets:
        _autosize(sheet)
        sheet.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
